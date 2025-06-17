from flask import Flask, render_template, redirect, url_for, request, session, flash, jsonify, send_from_directory
import os
import sys
import logging
import json
import datetime
from pathlib import Path

# Add project directory to path
sys.path.append(str(Path(__file__).resolve().parent))

# Import CRM-THABI modules
from src.clientes import GerenciadorClientes
from src.fornecedores import GerenciadorFornecedores
from src.produtos import GerenciadorProdutos
from src.vendas import GerenciadorVendas
from src.despesas import GerenciadorDespesas
from src.categorias import GerenciadorCategorias
from src.relatorios import GeradorRelatorios
from src.backup import BackupManager
from src.utils import formatar_cnpj, validar_cnpj, formatar_moeda
from src.margens import CalculadoraMargens
import threading

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Data directories
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
RELATORIOS_DIR = os.path.join(DATA_DIR, "relatorios")

# Create directories if they don't exist
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(RELATORIOS_DIR, exist_ok=True)

# Initialize managers
gerenciador_clientes = GerenciadorClientes(DATA_DIR)
gerenciador_fornecedores = GerenciadorFornecedores(DATA_DIR)
gerenciador_produtos = GerenciadorProdutos(DATA_DIR)
gerenciador_vendas = GerenciadorVendas(DATA_DIR)
gerenciador_despesas = GerenciadorDespesas(DATA_DIR)
gerenciador_categorias = GerenciadorCategorias(DATA_DIR)
gerador_relatorios = GeradorRelatorios(DATA_DIR, RELATORIOS_DIR)

# Initialize backup manager and start automatic backup
backup_manager = BackupManager(DATA_DIR)
# Run in a separate thread to avoid blocking
backup_thread = threading.Thread(target=backup_manager.iniciar_backup_automatico, args=(6,), daemon=True)
backup_thread.start()
logger.info("Backup automático configurado e iniciado")

# Create Flask application
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev_key_for_crm_thabi")

# Main routes - pages
@app.route('/')
def index():
    return redirect(url_for('dashboard'))

# Rota para servir arquivos estáticos da pasta data
@app.route('/data/<path:filename>')
def serve_data(filename):
    return send_from_directory(DATA_DIR, filename)

@app.route('/dashboard')
def dashboard():
    # Get data for dashboard
    clientes = gerenciador_clientes.obter_todos_clientes()
    vendas = gerenciador_vendas.obter_todas_vendas()
    despesas = gerenciador_despesas.obter_todas_despesas()
    
    # Calculate client statistics
    num_clientes = len(clientes)
    taxa_crescimento_clientes = 0  # Implement real calculation later
    
    # Calculate sales statistics
    valor_total_vendas = 0
    for venda in vendas:
        if venda.get('status_pagamento') != 'cancelado':
            valor_total_vendas += float(venda.get('valor', 0))
    taxa_crescimento_vendas = 0  # Implement real calculation later
    
    # Calculate expense statistics
    valor_total_despesas = 0
    for despesa in despesas:
        valor_total_despesas += float(despesa.get('valor', 0))
    taxa_crescimento_despesas = 0  # Implement real calculation later
    
    # Prepare recent sales (last 5)
    def parse_date_safe(date_str):
        """Tentativa de converter datas em diferentes formatos"""
        if not date_str:
            return datetime.datetime(2000, 1, 1)
            
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S']
        for fmt in formats:
            try:
                return datetime.datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        # Se nenhum formato funcionar, retorna uma data padrão    
        return datetime.datetime(2000, 1, 1)
        
    vendas_recentes = sorted(
        vendas, 
        key=lambda x: parse_date_safe(x.get('data_saida', '01/01/2000')), 
        reverse=True
    )[:5]
    
    # Process sales data for display
    for venda in vendas_recentes:
        # Add client name if possible
        cliente_id = venda.get('destinatario', '').split('/')[-1] if '/' in venda.get('destinatario', '') else None
        if cliente_id and cliente_id.isdigit():
            cliente = gerenciador_clientes.obter_cliente_por_id(int(cliente_id))
            if cliente:
                venda['cliente_nome'] = cliente['nome']
            else:
                venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
        else:
            venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
    
    # Prepare recent expenses (last 5)
    despesas_recentes = sorted(
        despesas, 
        key=lambda x: parse_date_safe(x.get('data', '01/01/2000')), 
        reverse=True
    )[:5]
    
    return render_template(
        'dashboard.html', 
        title="Dashboard - CRM THABI",
        num_clientes=num_clientes,
        taxa_crescimento_clientes=taxa_crescimento_clientes,
        valor_total_vendas=valor_total_vendas,
        taxa_crescimento_vendas=taxa_crescimento_vendas,
        valor_total_despesas=valor_total_despesas,
        taxa_crescimento_despesas=taxa_crescimento_despesas,
        vendas_recentes=vendas_recentes,
        despesas_recentes=despesas_recentes,
        clientes=clientes  # Added to show client list
    )

@app.route('/clientes')
def clientes():
    # Get all clients
    clientes = gerenciador_clientes.obter_todos_clientes()
    return render_template('clientes.html', title="Clientes - CRM THABI", clientes=clientes)

@app.route('/fornecedores')
def fornecedores():
    # Get all suppliers
    fornecedores = gerenciador_fornecedores.obter_todos_fornecedores()
    
    # Count products by supplier
    produtos_por_fornecedor = {}
    produtos = gerenciador_produtos.obter_todos_produtos()
    
    for produto in produtos:
        id_fornecedor = produto['id_fornecedor']
        if id_fornecedor in produtos_por_fornecedor:
            produtos_por_fornecedor[id_fornecedor] += 1
        else:
            produtos_por_fornecedor[id_fornecedor] = 1
    
    return render_template(
        'fornecedores.html', 
        title="Fornecedores - CRM THABI", 
        fornecedores=fornecedores,
        produtos_por_fornecedor=produtos_por_fornecedor
    )

@app.route('/produtos')
def produtos():
    # Get all products and suppliers
    produtos = gerenciador_produtos.obter_todos_produtos()
    fornecedores = gerenciador_fornecedores.obter_todos_fornecedores()
    
    # Add supplier information for each product
    for produto in produtos:
        fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(produto['id_fornecedor'])
        produto['fornecedor_nome'] = fornecedor['nome'] if fornecedor else "Desconhecido"
        produto['margem_lucro'] = gerenciador_produtos.calcular_margem_lucro(produto['id'])
    
    return render_template(
        'produtos.html', 
        title="Produtos - CRM THABI", 
        produtos=produtos,
        fornecedores=fornecedores
    )

@app.route('/vendas')
def vendas():
    # Get clients and products for sales form
    clientes = gerenciador_clientes.obter_todos_clientes()
    produtos = gerenciador_produtos.obter_todos_produtos()
    
    # Get all registered sales
    vendas = gerenciador_vendas.obter_todas_vendas()
    
    # Process sales data for display
    for venda in vendas:
        # Add client name if possible
        cliente_id = venda.get('destinatario', '').split('/')[-1] if '/' in venda.get('destinatario', '') else None
        if cliente_id and cliente_id.isdigit():
            cliente = gerenciador_clientes.obter_cliente_por_id(int(cliente_id))
            if cliente:
                venda['cliente_nome'] = cliente['nome']
            else:
                venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
        else:
            venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
    
    return render_template(
        'vendas.html', 
        title="Vendas - CRM THABI",
        vendas=vendas,
        clientes=clientes,
        produtos=produtos
    )

@app.route('/despesas')
def despesas():
    # Get suppliers for expense form
    fornecedores = gerenciador_fornecedores.obter_todos_fornecedores()
    
    # Get all categories
    categorias = gerenciador_categorias.obter_todas_categorias()
    
    # Get all registered expenses
    despesas = gerenciador_despesas.obter_todas_despesas()
    
    # Add supplier name to each expense
    for despesa in despesas:
        if despesa.get('fornecedor_id'):
            fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(despesa['fornecedor_id'])
            despesa['fornecedor_nome'] = fornecedor['nome'] if fornecedor else "Desconhecido"
        else:
            despesa['fornecedor_nome'] = "Não especificado"
    
    # Apply filters if there are parameters
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    categoria = request.args.get('categoria')
    fornecedor_id = request.args.get('fornecedor_id')
    
    # Filter by date
    if data_inicio and data_fim:
        try:
            # Convert to format dd/mm/yyyy
            data_inicio_obj = datetime.datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim_obj = datetime.datetime.strptime(data_fim, '%Y-%m-%d')
            data_inicio_str = data_inicio_obj.strftime('%d/%m/%Y')
            data_fim_str = data_fim_obj.strftime('%d/%m/%Y')
            
            despesas = gerenciador_despesas.obter_despesas_por_periodo(data_inicio_str, data_fim_str)
        except ValueError:
            flash('Invalid date format. Use the correct format.', 'error')
    
    # Filter by category
    if categoria:
        despesas = gerenciador_despesas.obter_despesas_por_categoria(categoria)
    
    # Filter by supplier
    if fornecedor_id and fornecedor_id.isdigit():
        despesas = gerenciador_despesas.obter_despesas_por_fornecedor(int(fornecedor_id))
    
    return render_template(
        'despesas.html', 
        title="Despesas - CRM THABI",
        despesas=despesas,
        fornecedores=fornecedores,
        categorias=categorias
    )

@app.route('/relatorios')
def relatorios():
    return render_template('relatorios.html', title="Relatórios - CRM THABI")

# API Endpoints - AJAX
@app.route('/api/clientes', methods=['GET', 'POST'])
def api_clientes():
    if request.method == 'GET':
        termo_busca = request.args.get('termo', '')
        if termo_busca:
            clientes = gerenciador_clientes.buscar_cliente(termo_busca)
        else:
            clientes = gerenciador_clientes.obter_todos_clientes()
        return jsonify(clientes)
    
    elif request.method == 'POST':
        data = request.json
        
        # Adicionar logs para depuração
        logger.info(f"POST /api/clientes - Dados recebidos: {data}")
        
        # Verificar se todos os campos estão presentes
        if not all(key in data for key in ['nome', 'numero_loja', 'cnpj']):
            logger.error("Campos obrigatórios ausentes na requisição")
            return jsonify({"success": False, "message": "Dados incompletos. Verifique todos os campos obrigatórios."}), 400
            
        resultado = gerenciador_clientes.adicionar_cliente(
            data['nome'], 
            data['numero_loja'], 
            data['cnpj'], 
            data.get('grupo', '')  # Tornar o grupo opcional
        )
        if resultado:
            return jsonify({"success": True, "message": "Cliente adicionado com sucesso!"}), 201
        else:
            return jsonify({"success": False, "message": "Erro ao adicionar cliente. Verifique os dados."}), 400

@app.route('/api/clientes/<int:id_cliente>', methods=['GET', 'PUT', 'DELETE'])
def api_cliente(id_cliente):
    if request.method == 'GET':
        cliente = gerenciador_clientes.obter_cliente_por_id(id_cliente)
        if cliente:
            return jsonify(cliente)
        return jsonify({"error": "Cliente não encontrado"}), 404
    
    elif request.method == 'PUT':
        data = request.json
        resultado = gerenciador_clientes.atualizar_cliente(
            id_cliente,
            data['nome'], 
            data['numero_loja'], 
            data['cnpj'], 
            data['grupo']
        )
        if resultado:
            return jsonify({"success": True, "message": "Cliente atualizado com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao atualizar cliente."}), 400
    
    elif request.method == 'DELETE':
        resultado = gerenciador_clientes.remover_cliente(id_cliente)
        if resultado:
            return jsonify({"success": True, "message": "Cliente removido com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao remover cliente."}), 400

@app.route('/api/fornecedores', methods=['GET', 'POST'])
def api_fornecedores():
    if request.method == 'GET':
        termo_busca = request.args.get('termo', '')
        if termo_busca:
            fornecedores = gerenciador_fornecedores.buscar_fornecedor(termo_busca)
        else:
            fornecedores = gerenciador_fornecedores.obter_todos_fornecedores()
        return jsonify(fornecedores)
    
    elif request.method == 'POST':
        data = request.json
        resultado = gerenciador_fornecedores.adicionar_fornecedor(
            data['nome'], 
            data['cnpj']
        )
        if resultado:
            return jsonify({"success": True, "message": "Fornecedor adicionado com sucesso!"}), 201
        else:
            return jsonify({"success": False, "message": "Erro ao adicionar fornecedor. Verifique os dados."}), 400

@app.route('/api/fornecedores/<int:id_fornecedor>', methods=['GET', 'PUT', 'DELETE'])
def api_fornecedor(id_fornecedor):
    if request.method == 'GET':
        fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(id_fornecedor)
        if fornecedor:
            return jsonify(fornecedor)
        return jsonify({"error": "Fornecedor não encontrado"}), 404
    
    elif request.method == 'PUT':
        data = request.json
        resultado = gerenciador_fornecedores.atualizar_fornecedor(
            id_fornecedor,
            data['nome'], 
            data['cnpj']
        )
        if resultado:
            return jsonify({"success": True, "message": "Fornecedor atualizado com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao atualizar fornecedor."}), 400
    
    elif request.method == 'DELETE':
        resultado = gerenciador_fornecedores.remover_fornecedor(id_fornecedor)
        if resultado:
            return jsonify({"success": True, "message": "Fornecedor removido com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao remover fornecedor."}), 400

@app.route('/api/produtos', methods=['GET', 'POST'])
def api_produtos():
    if request.method == 'GET':
        termo_busca = request.args.get('termo', '')
        fornecedor_id = request.args.get('fornecedor_id')
        
        if termo_busca:
            produtos = gerenciador_produtos.buscar_produto(termo_busca)
        elif fornecedor_id and fornecedor_id.isdigit():
            produtos = gerenciador_produtos.obter_produtos_por_fornecedor(int(fornecedor_id))
        else:
            produtos = gerenciador_produtos.obter_todos_produtos()
        
        for produto in produtos:
            fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(produto['id_fornecedor'])
            produto['fornecedor_nome'] = fornecedor['nome'] if fornecedor else "Desconhecido"
            produto['margem_lucro'] = gerenciador_produtos.calcular_margem_lucro(produto['id'])
            
        return jsonify(produtos)
    
    elif request.method == 'POST':
        data = request.json
        resultado = gerenciador_produtos.adicionar_produto(
            data['nome'], 
            float(data['valor_compra']), 
            float(data['valor_venda']), 
            int(data['id_fornecedor'])
        )
        if resultado:
            return jsonify({"success": True, "message": "Produto adicionado com sucesso!"}), 201
        else:
            return jsonify({"success": False, "message": "Erro ao adicionar produto. Verifique os dados."}), 400

@app.route('/api/produtos/<int:id_produto>', methods=['GET', 'PUT', 'DELETE'])
def api_produto(id_produto):
    if request.method == 'GET':
        produto = gerenciador_produtos.obter_produto_por_id(id_produto)
        if produto:
            fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(produto['id_fornecedor'])
            produto['fornecedor_nome'] = fornecedor['nome'] if fornecedor else "Desconhecido"
            produto['margem_lucro'] = gerenciador_produtos.calcular_margem_lucro(produto['id'])
            return jsonify(produto)
        return jsonify({"error": "Produto não encontrado"}), 404
    
    elif request.method == 'PUT':
        data = request.json
        resultado = gerenciador_produtos.atualizar_produto(
            id_produto,
            data['nome'], 
            float(data['valor_compra']), 
            float(data['valor_venda']), 
            int(data['id_fornecedor'])
        )
        if resultado:
            return jsonify({"success": True, "message": "Produto atualizado com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao atualizar produto."}), 400
    
    elif request.method == 'DELETE':
        resultado = gerenciador_produtos.remover_produto(id_produto)
        if resultado:
            return jsonify({"success": True, "message": "Produto removido com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao remover produto."}), 400

@app.route('/api/vendas', methods=['GET', 'POST'])
def api_vendas():
    if request.method == 'GET':
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        cliente_id = request.args.get('cliente_id')
        bonificacao = request.args.get('bonificacao')
        
        if data_inicio and data_fim:
            try:
                # Convert to format dd/mm/yyyy
                data_inicio_obj = datetime.datetime.strptime(data_inicio, '%Y-%m-%d')
                data_fim_obj = datetime.datetime.strptime(data_fim, '%Y-%m-%d')
                data_inicio_str = data_inicio_obj.strftime('%d/%m/%Y')
                data_fim_str = data_fim_obj.strftime('%d/%m/%Y')
                
                if bonificacao and bonificacao.lower() == 'true':
                    vendas = gerenciador_vendas.obter_bonificacoes_por_periodo(data_inicio_str, data_fim_str)
                else:
                    vendas = gerenciador_vendas.obter_vendas_por_periodo(data_inicio_str, data_fim_str)
            except ValueError:
                return jsonify({"error": "Formato de data inválido."}), 400
        elif cliente_id and cliente_id.isdigit():
            if bonificacao and bonificacao.lower() == 'true':
                vendas = gerenciador_vendas.obter_bonificacoes_por_cliente(id_cliente=int(cliente_id))
            else:
                vendas = gerenciador_vendas.obter_vendas_por_cliente(int(cliente_id))
        else:
            if bonificacao and bonificacao.lower() == 'true':
                vendas = [v for v in gerenciador_vendas.obter_todas_vendas() if v.get('bonificacao', False)]
            else:
                vendas = gerenciador_vendas.obter_todas_vendas()
        
        # Process sales data for display
        for venda in vendas:
            cliente_id = venda.get('destinatario', '').split('/')[-1] if '/' in venda.get('destinatario', '') else None
            if cliente_id and cliente_id.isdigit():
                cliente = gerenciador_clientes.obter_cliente_por_id(int(cliente_id))
                if cliente:
                    venda['cliente_nome'] = cliente['nome']
                else:
                    venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
            else:
                venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
                
        return jsonify(vendas)
    
    elif request.method == 'POST':
        data = request.json
        
        # Adicionar logs para depuração
        logger.info(f"POST /api/vendas - Dados recebidos: {data}")
        
        # Verificar se todos os campos obrigatórios estão presentes
        campos_obrigatorios = ['data_saida', 'valor', 'forma_pagamento']
        if not all(key in data for key in campos_obrigatorios):
            logger.error("Campos obrigatórios ausentes na requisição de venda")
            return jsonify({"success": False, "message": "Dados incompletos. Verifique os campos obrigatórios."}), 400
            
        # Process client reference
        if data.get('cliente_id'):
            destinatario = f"cliente/{data['cliente_id']}"
        else:
            destinatario = data.get('destinatario', 'Cliente Avulso')
        
        # Converter o valor para float, garantindo que está no formato correto
        try:
            # Tratar formatação no formato brasileiro (15.958,32 -> 15958.32)
            valor_str = str(data['valor']).replace('R$', '').strip()
            
            logger.info(f"Valor recebido para processamento: '{valor_str}'")
            
            # Verificar se o valor contém vírgula (formato brasileiro)
            if ',' in valor_str:
                # Remover todos os pontos (que são separadores de milhar) e substituir a vírgula por ponto
                valor_str = valor_str.replace('.', '')
                valor_str = valor_str.replace(',', '.')
                logger.info(f"Valor após conversão de formato brasileiro: '{valor_str}'")
            else:
                # Se não tem vírgula, pode ser que já esteja no formato americano (1595832.00)
                # ou que seja um número inteiro sem decimais
                logger.info(f"Valor mantido no formato original: '{valor_str}'")
            
            valor = float(valor_str)
            logger.info(f"Valor convertido para número: {valor} (de entrada: {data['valor']})")
        except (ValueError, TypeError) as e:
            logger.error(f"Erro ao converter valor: {e} - valor recebido: {data['valor']}")
            return jsonify({"success": False, "message": "Valor inválido. Forneça um número válido."}), 400
            
        # Agora a função adicionar_venda aceita status_pagamento e bonificacao como parâmetros opcionais
        resultado = gerenciador_vendas.adicionar_venda(
            data.get('numero_nota', ''),
            data['data_saida'],
            destinatario,
            valor,
            data['forma_pagamento'],
            data.get('data_vencimento', ''),  # Usando data_vencimento como data_pagar
            data.get('produtos', []),  # List of product IDs
            data.get('status_pagamento', 'pendente'),  # Status de pagamento explícito
            data.get('bonificacao', False)  # Indica se é uma bonificação
        )
        if resultado:
            return jsonify({"success": True, "message": "Venda adicionada com sucesso!"}), 201
        else:
            return jsonify({"success": False, "message": "Erro ao adicionar venda. Verifique os dados."}), 400

@app.route('/api/vendas/<int:id_venda>', methods=['GET', 'PUT', 'DELETE'])
def api_venda(id_venda):
    if request.method == 'GET':
        venda = gerenciador_vendas.obter_venda_por_id(id_venda)
        if venda:
            cliente_id = venda.get('destinatario', '').split('/')[-1] if '/' in venda.get('destinatario', '') else None
            if cliente_id and cliente_id.isdigit():
                cliente = gerenciador_clientes.obter_cliente_por_id(int(cliente_id))
                if cliente:
                    venda['cliente_nome'] = cliente['nome']
                else:
                    venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
            else:
                venda['cliente_nome'] = venda.get('destinatario', 'Desconhecido')
            return jsonify(venda)
        return jsonify({"error": "Venda não encontrada"}), 404
    
    elif request.method == 'PUT':
        data = request.json
        
        # Adicionar logs para depuração
        logger.info(f"PUT /api/vendas/{id_venda} - Dados recebidos: {data}")
        
        # Verificar se todos os campos obrigatórios estão presentes
        campos_obrigatorios = ['data_saida', 'valor', 'forma_pagamento']
        if not all(key in data for key in campos_obrigatorios):
            logger.error("Campos obrigatórios ausentes na requisição de atualização de venda")
            return jsonify({"success": False, "message": "Dados incompletos. Verifique os campos obrigatórios."}), 400
            
        # Process client reference
        if data.get('cliente_id'):
            destinatario = f"cliente/{data['cliente_id']}"
        else:
            destinatario = data.get('destinatario', 'Cliente Avulso')
            
        # Converter o valor para float, garantindo que está no formato correto
        try:
            # Tratar formatação no formato brasileiro (15.958,32 -> 15958.32)
            valor_str = str(data['valor']).replace('R$', '').strip()
            
            logger.info(f"Valor recebido para processamento na atualização: '{valor_str}'")
            
            # Verificar se o valor contém vírgula (formato brasileiro)
            if ',' in valor_str:
                # Remover todos os pontos (que são separadores de milhar) e substituir a vírgula por ponto
                valor_str = valor_str.replace('.', '')
                valor_str = valor_str.replace(',', '.')
                logger.info(f"Valor após conversão de formato brasileiro (atualização): '{valor_str}'")
            else:
                # Se não tem vírgula, pode ser que já esteja no formato americano (1595832.00)
                # ou que seja um número inteiro sem decimais
                logger.info(f"Valor mantido no formato original (atualização): '{valor_str}'")
            
            valor = float(valor_str)
            logger.info(f"Valor convertido para atualização: {valor} (de entrada: {data['valor']})")
        except (ValueError, TypeError) as e:
            logger.error(f"Erro ao converter valor para atualização: {e} - valor recebido: {data['valor']}")
            return jsonify({"success": False, "message": "Valor inválido. Forneça um número válido."}), 400
            
        # Observe que o método atualizar_venda define status_pagamento como parâmetro opcional antes de produtos
        resultado = gerenciador_vendas.atualizar_venda(
            id_venda,
            data['numero_nota'],
            data['data_saida'],
            destinatario,
            valor,
            data['forma_pagamento'],
            data.get('data_vencimento', ''),  # Corrigido para data_vencimento
            data.get('status_pagamento', 'pendente'),  # Status pagamento explícito
            data.get('produtos', []),  # List of product IDs
            data.get('bonificacao', False)  # Indica se é uma bonificação
        )
        if resultado:
            return jsonify({"success": True, "message": "Venda atualizada com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao atualizar venda."}), 400
    
    elif request.method == 'DELETE':
        resultado = gerenciador_vendas.remover_venda(id_venda)
        if resultado:
            return jsonify({"success": True, "message": "Venda removida com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao remover venda."}), 400

@app.route('/api/despesas', methods=['GET', 'POST'])
def api_despesas():
    if request.method == 'GET':
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        categoria = request.args.get('categoria')
        fornecedor_id = request.args.get('fornecedor_id')
        
        if data_inicio and data_fim:
            try:
                # Convert to format dd/mm/yyyy
                data_inicio_obj = datetime.datetime.strptime(data_inicio, '%Y-%m-%d')
                data_fim_obj = datetime.datetime.strptime(data_fim, '%Y-%m-%d')
                data_inicio_str = data_inicio_obj.strftime('%d/%m/%Y')
                data_fim_str = data_fim_obj.strftime('%d/%m/%Y')
                
                despesas = gerenciador_despesas.obter_despesas_por_periodo(data_inicio_str, data_fim_str)
            except ValueError:
                return jsonify({"error": "Formato de data inválido."}), 400
        elif categoria:
            despesas = gerenciador_despesas.obter_despesas_por_categoria(categoria)
        elif fornecedor_id and fornecedor_id.isdigit():
            despesas = gerenciador_despesas.obter_despesas_por_fornecedor(int(fornecedor_id))
        else:
            despesas = gerenciador_despesas.obter_todas_despesas()
        
        # Add supplier name to each expense
        for despesa in despesas:
            if despesa.get('fornecedor_id'):
                fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(despesa['fornecedor_id'])
                despesa['fornecedor_nome'] = fornecedor['nome'] if fornecedor else "Desconhecido"
            else:
                despesa['fornecedor_nome'] = "Não especificado"
                
        return jsonify(despesas)
    
    elif request.method == 'POST':
        data = request.json
        fornecedor_id = int(data['fornecedor_id']) if data.get('fornecedor_id') and data['fornecedor_id'].isdigit() else None
        
        resultado = gerenciador_despesas.adicionar_despesa(
            data['descricao'],
            float(data['valor']),
            data['data'],
            data['categoria'],
            fornecedor_id,
            data.get('numero_nota', ''),
            data.get('vencimento', ''),
            data.get('status', 'pendente')
        )
        if resultado:
            return jsonify({"success": True, "message": "Despesa adicionada com sucesso!"}), 201
        else:
            return jsonify({"success": False, "message": "Erro ao adicionar despesa. Verifique os dados."}), 400

@app.route('/api/despesas/<int:id_despesa>', methods=['GET', 'PUT', 'DELETE'])
def api_despesa(id_despesa):
    if request.method == 'GET':
        despesa = gerenciador_despesas.obter_despesa_por_id(id_despesa)
        if despesa:
            if despesa.get('fornecedor_id'):
                fornecedor = gerenciador_fornecedores.obter_fornecedor_por_id(despesa['fornecedor_id'])
                despesa['fornecedor_nome'] = fornecedor['nome'] if fornecedor else "Desconhecido"
            else:
                despesa['fornecedor_nome'] = "Não especificado"
            return jsonify(despesa)
        return jsonify({"error": "Despesa não encontrada"}), 404
    
    elif request.method == 'PUT':
        data = request.json
        fornecedor_id = int(data['fornecedor_id']) if data.get('fornecedor_id') and data['fornecedor_id'].isdigit() else None
        
        resultado = gerenciador_despesas.atualizar_despesa(
            id_despesa,
            data['descricao'],
            float(data['valor']),
            data['data'],
            data['categoria'],
            fornecedor_id,
            data.get('numero_nota', ''),
            data.get('vencimento', ''),
            data.get('status', 'pendente')
        )
        if resultado:
            return jsonify({"success": True, "message": "Despesa atualizada com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao atualizar despesa."}), 400
    
    elif request.method == 'DELETE':
        resultado = gerenciador_despesas.remover_despesa(id_despesa)
        if resultado:
            return jsonify({"success": True, "message": "Despesa removida com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao remover despesa."}), 400

@app.route('/api/categorias', methods=['GET', 'POST'])
def api_categorias():
    if request.method == 'GET':
        categorias = gerenciador_categorias.obter_todas_categorias()
        return jsonify(categorias)
    
    elif request.method == 'POST':
        data = request.json
        resultado = gerenciador_categorias.adicionar_categoria(data['nome'])
        if resultado:
            return jsonify({"success": True, "message": "Categoria adicionada com sucesso!"}), 201
        else:
            return jsonify({"success": False, "message": "Erro ao adicionar categoria. Verifique os dados."}), 400

@app.route('/api/categorias/<int:id_categoria>', methods=['GET', 'PUT', 'DELETE'])
def api_categoria(id_categoria):
    if request.method == 'GET':
        categoria = gerenciador_categorias.obter_categoria_por_id(id_categoria)
        if categoria:
            return jsonify(categoria)
        return jsonify({"error": "Categoria não encontrada"}), 404
    
    elif request.method == 'PUT':
        data = request.json
        resultado = gerenciador_categorias.atualizar_categoria(id_categoria, data['nome'])
        if resultado:
            return jsonify({"success": True, "message": "Categoria atualizada com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao atualizar categoria."}), 400
    
    elif request.method == 'DELETE':
        resultado = gerenciador_categorias.remover_categoria(id_categoria)
        if resultado:
            return jsonify({"success": True, "message": "Categoria removida com sucesso!"})
        return jsonify({"success": False, "message": "Erro ao remover categoria."}), 400

@app.route('/api/relatorios/<tipo>', methods=['GET'])
def api_relatorios(tipo):
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    if not data_inicio or not data_fim:
        return jsonify({"error": "Data inicial e final são obrigatórias."}), 400
    
    try:
        # Convert to format dd/mm/yyyy
        data_inicio_obj = datetime.datetime.strptime(data_inicio, '%Y-%m-%d')
        data_fim_obj = datetime.datetime.strptime(data_fim, '%Y-%m-%d')
        data_inicio_str = data_inicio_obj.strftime('%d/%m/%Y')
        data_fim_str = data_fim_obj.strftime('%d/%m/%Y')
    except ValueError:
        return jsonify({"error": "Formato de data inválido."}), 400
    
    if tipo == 'vendas':
        resultado = gerador_relatorios.gerar_relatorio_vendas(data_inicio_str, data_fim_str)
    elif tipo == 'produtos':
        resultado = gerador_relatorios.gerar_relatorio_produtos()
    elif tipo == 'despesas':
        resultado = gerador_relatorios.gerar_relatorio_despesas(data_inicio_str, data_fim_str)
    elif tipo == 'clientes':
        resultado = gerador_relatorios.gerar_relatorio_clientes(data_inicio_str, data_fim_str)
    elif tipo == 'lucro':
        resultado = gerador_relatorios.gerar_relatorio_lucro(data_inicio_str, data_fim_str)
    else:
        return jsonify({"error": "Tipo de relatório inválido."}), 400
    
    if resultado:
        return jsonify({"success": True, "relatorio": resultado})
    return jsonify({"success": False, "message": "Erro ao gerar relatório."}), 400

@app.route('/api/exportar/<tipo>', methods=['GET'])
def api_exportar(tipo):
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    formato = request.args.get('formato', 'xlsx')
    
    if not data_inicio or not data_fim:
        return jsonify({"error": "Data inicial e final são obrigatórias."}), 400
    
    try:
        # Convert to format dd/mm/yyyy
        data_inicio_obj = datetime.datetime.strptime(data_inicio, '%Y-%m-%d')
        data_fim_obj = datetime.datetime.strptime(data_fim, '%Y-%m-%d')
        data_inicio_str = data_inicio_obj.strftime('%d/%m/%Y')
        data_fim_str = data_fim_obj.strftime('%d/%m/%Y')
    except ValueError:
        return jsonify({"error": "Formato de data inválido."}), 400
    
    # Configura o diretório para salvar os relatórios
    gerador_relatorios.diretorio_relatorios = "data/relatorios"
    os.makedirs("data/relatorios", exist_ok=True)
    
    if tipo == 'vendas':
        caminho_arquivo = gerador_relatorios.exportar_relatorio_vendas(data_inicio_str, data_fim_str, formato)
    elif tipo == 'produtos':
        caminho_arquivo = gerador_relatorios.exportar_relatorio_produtos(formato)
    elif tipo == 'despesas':
        caminho_arquivo = gerador_relatorios.exportar_relatorio_despesas(data_inicio_str, data_fim_str, formato)
    elif tipo == 'clientes':
        caminho_arquivo = gerador_relatorios.exportar_relatorio_clientes(data_inicio_str, data_fim_str, formato)
    elif tipo == 'lucro':
        caminho_arquivo = gerador_relatorios.exportar_relatorio_lucro(data_inicio_str, data_fim_str, formato)
    else:
        return jsonify({"error": "Tipo de relatório inválido."}), 400
    
    if caminho_arquivo:
        nome_arquivo = os.path.basename(caminho_arquivo)
        return jsonify({"success": True, "arquivo": nome_arquivo, "caminho": caminho_arquivo})
    return jsonify({"success": False, "message": "Erro ao exportar relatório."}), 400

@app.route('/api/exportar_vendas_excel', methods=['GET'])
def api_exportar_vendas_excel():
    """Exporta todas as vendas para um arquivo Excel"""
    try:
        # Diretório onde o arquivo será salvo
        diretorio_exportacao = os.path.join(DATA_DIR, "exportacao")
        os.makedirs(diretorio_exportacao, exist_ok=True)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"vendas_exportadas_{timestamp}.xlsx"
        caminho_completo = os.path.join(diretorio_exportacao, nome_arquivo)
        
        # Obter todas as vendas
        vendas = gerenciador_vendas.obter_todas_vendas()
        
        # Exportar para Excel usando pandas
        import pandas as pd
        
        # Preparar dados para o DataFrame
        dados = []
        for venda in vendas:
            # Obter informações do cliente se disponível
            cliente_nome = "Cliente Avulso"
            numero_loja = ""
            if "cliente/" in venda.get('destinatario', ''):
                cliente_id = venda['destinatario'].split('/')[-1]
                cliente = gerenciador_clientes.obter_cliente_por_id(int(cliente_id))
                if cliente:
                    cliente_nome = cliente['nome']
                    numero_loja = cliente.get('numero_loja', '')
            
            # Formatar valor para exibição
            valor_formatado = f"R$ {float(venda['valor']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Flag de bonificação
            bonificacao = "Sim" if venda.get('bonificacao', False) else "Não"
            
            dados.append({
                'ID': venda['id'],
                'Número Nota': venda.get('numero_nota', ''),
                'Data': venda['data_saida'],
                'Cliente': cliente_nome,
                'Número Loja': numero_loja,
                'Valor': valor_formatado,
                'Forma Pagamento': venda['forma_pagamento'],
                'Data Vencimento': venda.get('data_pagar', ''),
                'Status': venda['status_pagamento'],
                'Bonificação': bonificacao
            })
        
        # Criar DataFrame e exportar
        df = pd.DataFrame(dados)
        df.to_excel(caminho_completo, index=False, sheet_name='Vendas')
        
        # Ajustar largura das colunas
        with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Vendas')
            worksheet = writer.sheets['Vendas']
            for i, coluna in enumerate(df.columns):
                largura_coluna = max(df[coluna].astype(str).map(len).max(), len(coluna)) + 2
                worksheet.column_dimensions[chr(65 + i)].width = largura_coluna
        
        return jsonify({
            "success": True, 
            "message": "Vendas exportadas com sucesso!", 
            "arquivo": nome_arquivo,
            "caminho": caminho_completo
        })
    
    except Exception as e:
        logger.error(f"Erro ao exportar vendas para Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Erro ao exportar vendas: {str(e)}"}), 500

@app.route('/api/gerar_modelo_importacao', methods=['GET'])
def api_gerar_modelo_importacao():
    """Gera um modelo de planilha Excel para importação de vendas"""
    try:
        # Diretório onde o arquivo será salvo
        diretorio_modelos = os.path.join(DATA_DIR, "modelos")
        os.makedirs(diretorio_modelos, exist_ok=True)
        
        # Nome do arquivo
        nome_arquivo = "modelo_importacao_vendas.xlsx"
        caminho_completo = os.path.join(diretorio_modelos, nome_arquivo)
        
        # Criar modelo usando pandas
        import pandas as pd
        
        # Colunas do modelo
        colunas = [
            'Número Nota', 
            'Data', 
            'Cliente ID', 
            'Cliente Nome (para avulso)', 
            'Valor',
            'Forma Pagamento', 
            'Data Vencimento', 
            'Status'
        ]
        
        # Dados de exemplo
        dados_exemplo = [
            {
                'Número Nota': '12345', 
                'Data': '31/05/2025', 
                'Cliente ID': '1', 
                'Cliente Nome (para avulso)': '', 
                'Valor': 'R$ 1.500,00',
                'Forma Pagamento': 'boleto', 
                'Data Vencimento': '15/06/2025', 
                'Status': 'pendente'
            },
            {
                'Número Nota': '12346', 
                'Data': '31/05/2025', 
                'Cliente ID': '', 
                'Cliente Nome (para avulso)': 'João da Silva', 
                'Valor': 'R$ 2.750,00',
                'Forma Pagamento': 'pix', 
                'Data Vencimento': '', 
                'Status': 'pago'
            },
        ]
        
        # Criar DataFrame
        df = pd.DataFrame(dados_exemplo)
        
        # Adicionar uma folha de instruções
        with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Modelo')
            
            # Ajustar largura das colunas
            worksheet = writer.sheets['Modelo']
            for i, coluna in enumerate(df.columns):
                largura_coluna = max(df[coluna].astype(str).map(len).max(), len(coluna)) + 2
                worksheet.column_dimensions[chr(65 + i)].width = largura_coluna
            
            # Criar folha de instruções
            instrucoes_df = pd.DataFrame({
                'Instruções': [
                    'Como usar este modelo para importar vendas:',
                    '',
                    '1. Preencha os dados conforme os exemplos fornecidos',
                    '2. Cliente ID: use o ID numérico do cliente cadastrado no sistema',
                    '3. Cliente Nome: preencha apenas se for venda para cliente avulso (sem cadastro)',
                    '4. Valor: formato monetário brasileiro (R$ 1.234,56)',
                    '5. Data: formato DD/MM/AAAA',
                    '6. Forma Pagamento: pix, dinheiro, cartao, boleto, cheque',
                    '7. Status: pendente, pago, atrasado, cancelado',
                    '',
                    'Observações:',
                    '- A coluna "Data Vencimento" é obrigatória apenas para formas de pagamento a prazo',
                    '- Para clientes cadastrados, preencha apenas a coluna "Cliente ID" e deixe "Cliente Nome" em branco',
                    '- Para clientes avulsos, deixe "Cliente ID" em branco e preencha "Cliente Nome"'
                ]
            })
            instrucoes_df.to_excel(writer, index=False, sheet_name='Instruções')
            
            # Ajustar largura da coluna de instruções
            worksheet = writer.sheets['Instruções']
            worksheet.column_dimensions['A'].width = 100
        
        return jsonify({
            "success": True, 
            "message": "Modelo de importação gerado com sucesso!", 
            "arquivo": nome_arquivo,
            "caminho": caminho_completo
        })
    
    except Exception as e:
        logger.error(f"Erro ao gerar modelo de importação: {str(e)}")
        return jsonify({"success": False, "message": f"Erro ao gerar modelo: {str(e)}"}), 500

@app.route('/api/exportar_vendas_por_ano', methods=['GET'])
def api_exportar_vendas_por_ano():
    """Exporta vendas do ano atual para um arquivo Excel com abas mensais"""
    try:
        # Diretório onde o arquivo será salvo
        diretorio_exportacao = os.path.join(DATA_DIR, "exportacao")
        os.makedirs(diretorio_exportacao, exist_ok=True)
        
        # Obter o ano atual (ou do parâmetro)
        ano = request.args.get('ano', datetime.datetime.now().year)
        
        # Nome do arquivo
        nome_arquivo = f"VENDAS_{ano}.xlsx"
        caminho_completo = os.path.join(diretorio_exportacao, nome_arquivo)
        
        # Obter todas as vendas
        todas_vendas = gerenciador_vendas.obter_todas_vendas()
        
        # Importar bibliotecas
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Criar um workbook vazio
        wb = Workbook()
        
        # Remover a planilha default
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Definir nomes dos meses e suas abreviações
        meses = [
            {"nome": "Janeiro", "abrev": "JAN", "num": "01"},
            {"nome": "Fevereiro", "abrev": "FEV", "num": "02"},
            {"nome": "Março", "abrev": "MAR", "num": "03"},
            {"nome": "Abril", "abrev": "ABR", "num": "04"},
            {"nome": "Maio", "abrev": "MAI", "num": "05"},
            {"nome": "Junho", "abrev": "JUN", "num": "06"},
            {"nome": "Julho", "abrev": "JUL", "num": "07"},
            {"nome": "Agosto", "abrev": "AGO", "num": "08"},
            {"nome": "Setembro", "abrev": "SET", "num": "09"},
            {"nome": "Outubro", "abrev": "OUT", "num": "10"},
            {"nome": "Novembro", "abrev": "NOV", "num": "11"},
            {"nome": "Dezembro", "abrev": "DEZ", "num": "12"}
        ]
        
        # Função para estilizar cabeçalhos
        def estilizar_cabecalho(ws):
            # Definir estilo para cabeçalhos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Aplicar estilo nos cabeçalhos
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Criar uma folha para cada mês
        for mes in meses:
            # Criar folha com nome abreviado do mês
            ws = wb.create_sheet(mes["abrev"])
            
            # Filtrar vendas do mês/ano especificado
            vendas_mes = []
            for venda in todas_vendas:
                try:
                    # Extrair mês e ano da data da venda
                    data_parts = venda['data_saida'].split('/')
                    # Formato esperado: DD/MM/AAAA
                    if len(data_parts) == 3:
                        venda_mes = data_parts[1]
                        venda_ano = data_parts[2]
                        
                        if venda_mes == mes["num"] and venda_ano == str(ano):
                            # Obter informações do cliente
                            cliente_nome = "Cliente Avulso"
                            numero_loja = ""
                            if "cliente/" in venda.get('destinatario', ''):
                                cliente_id = venda['destinatario'].split('/')[-1]
                                cliente = gerenciador_clientes.obter_cliente_por_id(int(cliente_id))
                                if cliente:
                                    cliente_nome = cliente['nome']
                                    numero_loja = cliente.get('numero_loja', '')
                                    
                            # Formatar valor
                            valor_formatado = f"R$ {float(venda['valor']):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                            
                            # Flag de bonificação
                            bonificacao = "Sim" if venda.get('bonificacao', False) else "Não"
                            
                            vendas_mes.append({
                                'ID': venda['id'],
                                'Número Nota': venda.get('numero_nota', ''),
                                'Data': venda['data_saida'],
                                'Cliente': cliente_nome,
                                'Número Loja': numero_loja,
                                'Valor': valor_formatado,
                                'Forma Pagamento': venda['forma_pagamento'],
                                'Data Vencimento': venda.get('data_pagar', ''),
                                'Status': venda['status_pagamento'],
                                'Bonificação': bonificacao
                            })
                except Exception as e:
                    logger.error(f"Erro ao processar venda para exportação: {str(e)}")
                    continue
            
            # Se houver vendas no mês, criar o DataFrame e adicionar à planilha
            if vendas_mes:
                df = pd.DataFrame(vendas_mes)
                
                # Adicionar o DataFrame à planilha
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Estilizar a planilha
                estilizar_cabecalho(ws)
                
                # Ajustar largura das colunas
                for i, coluna in enumerate(df.columns):
                    col_letter = chr(65 + i)
                    # Encontrar a largura máxima baseada no conteúdo
                    largura_max = max(len(str(x)) for x in df[coluna]) + 2
                    largura_coluna = max(largura_max, len(coluna)) + 2
                    ws.column_dimensions[col_letter].width = largura_coluna
            else:
                # Se não houver vendas, apenas adicionar cabeçalhos
                cabecalhos = ['ID', 'Número Nota', 'Data', 'Cliente', 'Número Loja', 'Valor', 
                             'Forma Pagamento', 'Data Vencimento', 'Status', 'Bonificação']
                for i, header in enumerate(cabecalhos, 1):
                    ws.cell(row=1, column=i, value=header)
                
                # Adicionar mensagem de "Sem dados"
                ws.cell(row=2, column=1, value="Sem vendas registradas neste mês")
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cabecalhos))
                
                # Estilizar cabeçalhos
                estilizar_cabecalho(ws)
                
                # Ajustar largura das colunas
                for i, header in enumerate(cabecalhos):
                    col_letter = chr(65 + i)
                    ws.column_dimensions[col_letter].width = len(header) + 4
        
        # Salvar o arquivo
        wb.save(caminho_completo)
        
        return jsonify({
            "success": True, 
            "message": f"Vendas do ano {ano} exportadas com sucesso!", 
            "arquivo": nome_arquivo,
            "caminho": caminho_completo
        })
    
    except Exception as e:
        logger.error(f"Erro ao exportar vendas por ano para Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Erro ao exportar vendas: {str(e)}"}), 500

@app.route('/api/importar_vendas_excel', methods=['POST'])
def api_importar_vendas_excel():
    """Importa vendas de um arquivo Excel"""
    if 'arquivo' not in request.files:
        return jsonify({"success": False, "message": "Nenhum arquivo enviado."}), 400
    
    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({"success": False, "message": "Nome de arquivo vazio."}), 400
    
    if not arquivo.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "Formato de arquivo inválido. Use .xlsx ou .xls."}), 400
    
    try:
        # Salvar arquivo temporariamente
        import tempfile
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, arquivo.filename)
        arquivo.save(temp_path)
        
        # Importar arquivo Excel
        import pandas as pd
        df = pd.read_excel(temp_path)
        
        # Validar estrutura do arquivo
        colunas_requeridas = [
            'Número Nota', 'Data', 'Valor', 'Forma Pagamento', 'Status'
        ]
        
        colunas_faltantes = [col for col in colunas_requeridas if col not in df.columns]
        if colunas_faltantes:
            return jsonify({
                "success": False, 
                "message": f"Colunas obrigatórias ausentes: {', '.join(colunas_faltantes)}"
            }), 400
            
        # Processar vendas
        vendas_importadas = 0
        vendas_com_erro = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                # Extrair dados básicos da linha
                numero_nota = str(row['Número Nota']) if not pd.isna(row['Número Nota']) else ''
                data_saida = str(row['Data']) if not pd.isna(row['Data']) else ''
                
                # Processar valor (remover formatação R$)
                valor_str = str(row['Valor']) if not pd.isna(row['Valor']) else '0'
                valor_str = valor_str.replace('R$', '').strip()
                # Verificar se o valor contém vírgula (formato brasileiro)
                if ',' in valor_str:
                    # Remover todos os pontos (que são separadores de milhar) e substituir a vírgula por ponto
                    valor_str = valor_str.replace('.', '')
                    valor_str = valor_str.replace(',', '.')
                try:
                    valor = float(valor_str)
                except ValueError:
                    erros.append(f"Linha {index+2}: Valor inválido '{row['Valor']}'")
                    vendas_com_erro += 1
                    continue
                
                forma_pagamento = str(row['Forma Pagamento']) if not pd.isna(row['Forma Pagamento']) else ''
                data_vencimento = str(row['Data Vencimento']) if 'Data Vencimento' in df.columns and not pd.isna(row['Data Vencimento']) else ''
                status = str(row['Status']) if not pd.isna(row['Status']) else ''
                
                # Identificar cliente ou destinatário
                cliente_id = None
                destinatario = "Cliente Avulso"
                
                if 'Cliente ID' in df.columns and not pd.isna(row['Cliente ID']) and str(row['Cliente ID']).strip():
                    cliente_id = int(row['Cliente ID'])
                    # Verificar se cliente existe
                    cliente = gerenciador_clientes.obter_cliente_por_id(cliente_id)
                    if not cliente:
                        erros.append(f"Linha {index+2}: Cliente ID {cliente_id} não encontrado")
                        vendas_com_erro += 1
                        continue
                    destinatario = f"cliente/{cliente_id}"
                elif 'Cliente Nome (para avulso)' in df.columns and not pd.isna(row['Cliente Nome (para avulso)']):
                    destinatario = str(row['Cliente Nome (para avulso)'])
                
                # Adicionar a venda
                resultado = gerenciador_vendas.adicionar_venda(
                    numero_nota,
                    data_saida,
                    destinatario,
                    valor,
                    forma_pagamento,
                    data_vencimento,
                    [],  # Sem produtos
                    status
                )
                
                if resultado:
                    vendas_importadas += 1
                else:
                    erros.append(f"Linha {index+2}: Erro ao adicionar venda")
                    vendas_com_erro += 1
            
            except Exception as e:
                erros.append(f"Linha {index+2}: {str(e)}")
                vendas_com_erro += 1
        
        # Limpar arquivos temporários
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        return jsonify({
            "success": True,
            "message": f"{vendas_importadas} vendas importadas com sucesso. {vendas_com_erro} vendas com erro.",
            "vendas_importadas": vendas_importadas,
            "vendas_com_erro": vendas_com_erro,
            "erros": erros
        })
    
    except Exception as e:
        logger.error(f"Erro ao importar vendas do Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Erro ao importar vendas: {str(e)}"}), 500

# API para calculadora de margens
@app.route('/api/calculadora/margem', methods=['POST'])
def api_calcular_margem():
    """Calcula a margem de lucro percentual com base no preço de venda e custo"""
    try:
        data = request.json
        preco_venda = float(data.get('preco_venda', 0))
        custo = float(data.get('custo', 0))
        
        if preco_venda <= 0 or custo <= 0:
            return jsonify({
                'success': False,
                'message': 'Preço de venda e custo devem ser maiores que zero'
            }), 400
            
        margem = CalculadoraMargens.calcular_margem_porcentagem(preco_venda, custo)
        
        return jsonify({
            'success': True,
            'margem': margem,
            'lucro': preco_venda - custo
        })
    except Exception as e:
        logger.error(f"Erro ao calcular margem: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao calcular: {str(e)}'
        }), 400

@app.route('/api/calculadora/preco-venda', methods=['POST'])
def api_calcular_preco_venda():
    """Calcula o preço de venda com base no custo e na margem desejada"""
    try:
        data = request.json
        custo = float(data.get('custo', 0))
        margem = float(data.get('margem', 0))
        
        if custo <= 0 or margem <= 0 or margem >= 100:
            return jsonify({
                'success': False,
                'message': 'Custo deve ser maior que zero e margem deve estar entre 0 e 100'
            }), 400
            
        preco_venda = CalculadoraMargens.calcular_preco_venda_por_margem(custo, margem)
        
        return jsonify({
            'success': True,
            'preco_venda': preco_venda,
            'lucro': preco_venda - custo
        })
    except Exception as e:
        logger.error(f"Erro ao calcular preço de venda: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao calcular: {str(e)}'
        }), 400

@app.route('/api/calculadora/preco-markup', methods=['POST'])
def api_calcular_preco_markup():
    """Calcula o preço de venda com base no custo e no markup (multiplicador)"""
    try:
        data = request.json
        custo = float(data.get('custo', 0))
        markup = float(data.get('markup', 0))
        
        if custo <= 0 or markup <= 0:
            return jsonify({
                'success': False,
                'message': 'Custo e markup devem ser maiores que zero'
            }), 400
            
        preco_venda = CalculadoraMargens.calcular_preco_venda_por_markup(custo, markup)
        
        return jsonify({
            'success': True,
            'preco_venda': preco_venda,
            'lucro': preco_venda - custo,
            'margem': CalculadoraMargens.calcular_margem_porcentagem(preco_venda, custo)
        })
    except Exception as e:
        logger.error(f"Erro ao calcular preço por markup: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao calcular: {str(e)}'
        }), 400

@app.route('/api/calculadora/analise', methods=['POST'])
def api_analisar_produto():
    """Realiza uma análise completa de um produto"""
    try:
        data = request.json
        custo = float(data.get('custo', 0))
        preco_venda = float(data.get('preco_venda', 0))
        quantidade = int(data.get('quantidade', 1))
        impostos_percentual = float(data.get('impostos', 0))
        custos_fixos = float(data.get('custos_fixos', 0))
        
        if custo <= 0 or preco_venda <= 0 or quantidade <= 0:
            return jsonify({
                'success': False,
                'message': 'Custo, preço de venda e quantidade devem ser maiores que zero'
            }), 400
            
        resultado = CalculadoraMargens.analisar_produto(
            custo, preco_venda, quantidade, impostos_percentual, custos_fixos
        )
        
        return jsonify({
            'success': True,
            'resultado': resultado
        })
    except Exception as e:
        logger.error(f"Erro na análise do produto: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao analisar: {str(e)}'
        }), 400

# Rota para a página da calculadora de margens
@app.route('/calculadora')
def calculadora_margens():
    return render_template('calculadora.html', title="Calculadora de Margens - CRM THABI")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
